import sys
import cv2
import numpy as np
import serial  # For serial communication
from serial import SerialException
from PyQt6.QtWidgets import QApplication, QMainWindow, QLabel, QPushButton, QComboBox, QVBoxLayout, QHBoxLayout, QWidget, QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox, QProgressDialog
from PyQt6.QtGui import QImage, QPixmap, QIcon, QFont
from PyQt6.QtCore import QTimer, Qt
import openpyxl
from datetime import datetime

class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()

        self.setWindowTitle("Machine Dashboard")
        self.setGeometry(100, 100, 1600, 900)

        main_layout = QHBoxLayout()

        left_layout = QVBoxLayout()
        left_layout.setContentsMargins(10, 10, 10, 10)

        self.data_table = QTableWidget(24, 6, self)
        self.data_table.setHorizontalHeaderLabels(["LOT ID", "CBD", "Maker", "BMS", "Total", "Timestamp"])
        
        self.data_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.data_table.setColumnWidth(0, int(self.data_table.columnWidth(0) * 1.5))
        left_layout.addWidget(self.data_table)

        button_layout = QHBoxLayout()
        self.upload_btn = QPushButton("Upload", self)
        self.upload_btn.setFixedSize(150, 50)
        self.upload_btn.setFont(QFont('Arial', 14))
        self.upload_btn.clicked.connect(self.export_to_excel)
        button_layout.addWidget(self.upload_btn, alignment=Qt.AlignmentFlag.AlignCenter)

        self.delete_btn = QPushButton("Delete", self)
        self.delete_btn.setFixedSize(150, 50)
        self.delete_btn.setFont(QFont('Arial', 14))
        self.delete_btn.clicked.connect(self.delete_data)
        button_layout.addWidget(self.delete_btn, alignment=Qt.AlignmentFlag.AlignCenter)

        left_layout.addLayout(button_layout)

        center_layout = QVBoxLayout()
        center_layout.setContentsMargins(10, 10, 10, 10)

        self.camera_label = QLabel(self)
        self.camera_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.camera_label.setStyleSheet("background-color: black;")
        center_layout.addWidget(self.camera_label)

        self.cam_select = QComboBox(self)
        self.cam_select.addItems(["Camera 1", "Camera 2"])
        center_layout.addWidget(self.cam_select)

        self.info_table = QTableWidget(4, 2, self)
        self.info_table.setHorizontalHeaderLabels(["Field", "Value"])
        self.info_table.setItem(0, 0, QTableWidgetItem("LOT ID"))
        self.info_table.setItem(1, 0, QTableWidgetItem("CBD"))
        self.info_table.setItem(2, 0, QTableWidgetItem("Maker"))
        self.info_table.setItem(3, 0, QTableWidgetItem("BMS"))
        self.info_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        center_layout.addWidget(self.info_table)

        self.confirm_btn = QPushButton("Confirm", self)
        self.confirm_btn.setFixedSize(150, 50)
        self.confirm_btn.setFont(QFont('Arial', 14))
        self.confirm_btn.clicked.connect(self.confirm_data)

        self.rescan_btn = QPushButton("Rescan", self)
        self.rescan_btn.setFixedSize(150, 50)
        self.rescan_btn.setFont(QFont('Arial', 14))
        self.rescan_btn.clicked.connect(self.rescan_data)

        right_button_layout = QHBoxLayout()
        right_button_layout.addWidget(self.rescan_btn, alignment=Qt.AlignmentFlag.AlignCenter)  
        right_button_layout.addWidget(self.confirm_btn, alignment=Qt.AlignmentFlag.AlignCenter)

        center_layout.addLayout(right_button_layout)

        center_layout.setStretch(0, 2)
        center_layout.setStretch(2, 1)

        main_layout.addLayout(left_layout)
        main_layout.addLayout(center_layout)

        main_layout.setStretch(0, 2)
        main_layout.setStretch(1, 3)

        widget = QWidget()
        widget.setLayout(main_layout)
        self.setCentralWidget(widget)

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_frame)
        self.capture = None

        self.cam_select.currentIndexChanged.connect(self.start_camera)

        # Serial setup
        self.serial_port = None
        self.connect_to_serial()

        # Timer to check for serial data
        self.serial_timer = QTimer(self)
        self.serial_timer.timeout.connect(self.read_serial_data)
        self.serial_timer.start(100)  # Check serial data every 100ms

        # Start camera and check connection
        self.connect_camera()

    def connect_to_serial(self):
        try:
            self.serial_port = serial.Serial('COM4', 115200, timeout=1)  # Replace 'COM4' with your actual port
            QMessageBox.information(self, "Microcontroller Connected", "ESP is connected. Running with ESP.")
        except SerialException:
            QMessageBox.warning(self, "Microcontroller Disconnected", "Microcontroller is disconnected. Running without ESP.")

    def read_serial_data(self):
        if self.serial_port:
            try:
                if self.serial_port.in_waiting > 0:
                    command = self.serial_port.readline().decode('utf-8').strip()
                    print(f"Received command: {command}")

                    if command == "upload":
                        self.export_to_excel()
                    elif command == "delete":
                        self.delete_data()
                    elif command == "rescan":
                        self.rescan_data()
                    elif command == "confirm":
                        self.confirm_data()
            except SerialException:
                QMessageBox.warning(self, "Microcontroller Disconnected", "Microcontroller is disconnected. Running without ESP.")
                self.serial_port = None  # Set to None to stop further checks

    def confirm_data(self):
        # Check if left table is full
        if self.data_table.rowCount() == 24 and self.data_table.item(23, 0) is not None:
            QMessageBox.warning(self, "Limit Reached", "Please upload the data before adding more.")
            return

        # Find the first empty row in the left table
        next_empty_row = None
        for row in range(self.data_table.rowCount()):
            if self.data_table.item(row, 0) is None:
                next_empty_row = row
                break
        
        if next_empty_row is None:
            QMessageBox.warning(self, "Limit Reached", "Please upload the data before adding more.")
            return

        # Transfer data from the right table to the left table
        data_added = False
        for row in range(self.info_table.rowCount()):
            item = self.info_table.item(row, 1)  # Get the item from the 'Value' column of the right table
            if item and item.text().strip():
                self.data_table.setItem(next_empty_row, row, QTableWidgetItem(item.text()))
                data_added = True

        # Add timestamp to the last column if data was added
        if data_added:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.data_table.setItem(next_empty_row, 5, QTableWidgetItem(timestamp))
            self.clear_right_table()

        # Force an immediate update of the GUI
        QApplication.processEvents()
    
    def clear_right_table(self):
        # Clear all items in the 'Value' column of the right table
        for row in range(self.info_table.rowCount()):
            self.info_table.setItem(row, 1, QTableWidgetItem(''))
    
    def rescan_data(self):
        # Clear the right table value column
        self.clear_right_table()

    def delete_data(self):
        # Clear all items in the data_table
        self.data_table.clearContents()
        print("Left table data deleted.")

    def export_to_excel(self):
        # Check if there is data to export
        if self.data_table.item(0, 0) is None:
            QMessageBox.warning(self, "No Data", "There is no data to export to Excel.")
            return

        file_path = "F:\\Project\\GUI\\Database_test.xlsx"
        
        try:
            # Load the existing workbook
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            # Find the next available row in the sheet
            next_row = sheet.max_row + 2

            # Write data from the table to the Excel sheet, starting from column 2 (B) to match "LOT ID"
            for row in range(self.data_table.rowCount()):
                if self.data_table.item(row, 0) is None:
                    break
                for col in range(self.data_table.columnCount()):
                    item = self.data_table.item(row, col)
                    if item is not None:
                        sheet.cell(row=next_row + row, column=col + 2, value=item.text())  # column=col + 2 to skip the "No." column

            # Save the workbook
            workbook.save(file_path)
            print(f"Data appended to {file_path}")
            
            # Clear the data on the dashboard
            self.clear_dashboard_data()

        except PermissionError:
            QMessageBox.warning(self, "File Open Error", "Please close the Excel file before uploading.")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"An error occurred: {e}")

    def clear_dashboard_data(self):
        self.data_table.clearContents()
        print("Dashboard data cleared.")

    def start_camera(self):
        if self.capture and self.capture.isOpened():
            self.capture.release()

        cam_index = self.cam_select.currentIndex()
        self.capture = cv2.VideoCapture(cam_index)

        if not self.capture.isOpened():
            print(f"Error: Unable to open camera {cam_index}")
            return

        self.capture.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
        self.capture.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)

        self.timer.start(30)

    def update_frame(self):
        if self.capture and self.capture.isOpened():
            ret, frame = self.capture.read()
            if ret:
                frame = cv2.flip(frame, 1)
                image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                height, width, channel = image.shape
                step = channel * width
                q_image = QImage(image.data, width, height, step, QImage.Format.Format_RGB888)
                
                self.camera_label.setPixmap(QPixmap.fromImage(q_image).scaled(
                    self.camera_label.size(), Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))

    def connect_camera(self):
        progress_dialog = QProgressDialog("Connecting to camera...", None, 0, 0, self)
        progress_dialog.setWindowTitle("Please wait")
        progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
        progress_dialog.show()

        # Try to connect to the default camera
        self.start_camera()

        # Wait for camera to connect and display
        while self.capture is None or not self.capture.isOpened():
            QApplication.processEvents()
            self.start_camera()
        
        progress_dialog.close()
        self.update_frame()

    def resizeEvent(self, event):
        if self.capture:
            self.update_frame()
        super(MainWindow, self).resizeEvent(event)

    def closeEvent(self, event):
        if self.capture:
            self.capture.release()

        if self.serial_port and self.serial_port.is_open:
            self.serial_port.close()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
